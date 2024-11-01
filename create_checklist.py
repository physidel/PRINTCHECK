#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PRINTCHECK - Automated Checklist Generator for STL Files in 3D Printing

This script generates a structured checklist for STL files, creating 3D previews
for each file and organizing them in an Excel sheet. It helps users ensure all
necessary files are reviewed and tracked before starting a 3D printing project.

Features:
- Automatic creation of an Excel checklist based on STL files.
- 3D previews for each STL file, color-coded based on file naming conventions.
- Organized output by folder and subfolder structure.
- Alerts for STL files that fail to generate previews, included in the checklist.

Color-coding conventions:
- Red: Files containing '[a]' in the name.
- White: Files containing '[c]' in the name.
- Black: All other files.

MIT License

-------------------------------------------------------------------------------
Revision History:
- v1.0.0 (2024-10-26): Initial release
    - Added automatic checklist generation with 3D previews.
    - Implemented color-coding based on file name patterns.
    - Provided Excel output with structured folder organization.

- v2.0.0 (2024-10-27): Updates and improvements
    - Added support for entering the STL folder path via command line or manual input.
    - Improved logging: log messages are collected and written to a timestamped log file.
    - Created a "logs" directory for storing log files.
    - Changed to use a temporary directory for storing STL previews, which are deleted after execution.
    - Added progress bar using tqdm for better user feedback during STL file processing.
    - Adding retries to create_3d_preview() to make it more stable, as the backend pyglet can fail occasionally.
-------------------------------------------------------------------------------
"""

import os
import sys
from pathlib import Path
import tempfile
import shutil
import trimesh
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from tqdm import tqdm
from datetime import datetime
import time

# Create logs directory if it doesn't exist
logs_dir = Path("logs")
logs_dir.mkdir(exist_ok=True)

# Check if the user provided the STL directory as a command-line argument
if len(sys.argv) > 1:
    stls_dir = Path(sys.argv[1])
else:
    # Prompt the user to enter the path if not provided
    stls_dir = Path(input("Please enter the path to the folder containing the STL files: "))

# Check if the provided path is valid
if not stls_dir.is_dir():
    print(f"The specified path '{stls_dir}' is not a valid directory. Exiting.")
    sys.exit(1)

# Set up a temporary directory for the STL previews
temp_dir = Path(tempfile.mkdtemp())
log_messages = []  # List to collect log messages
log_messages.append(f"Using temporary directory for previews: {temp_dir}")

# Gather all STL files in the specified 'Stls' directory and its subdirectories
stl_files = []
for root, _, files in os.walk(stls_dir):
    for file in files:
        if file.lower().endswith('.stl'):
            full_path = Path(root) / file
            stl_files.append(full_path)

log_messages.append(f"Found {len(stl_files)} STL files in '{stls_dir}'.")
print(f"Processing {len(stl_files)} STL files...")

def create_3d_preview(stl_file_path, save_path, image_size=(200, 200), max_retries=10):
    """
    Generate a 3D preview image for an STL file with auto-scaling and color coding.
    
    The function attempts to create a 3D preview image of the given STL file. It applies color coding 
    based on the file name, calculates the optimal camera settings for the preview, and saves the 
    rendered image to the specified path. It includes a retry mechanism to handle transient failures 
    during the rendering process.
    
    Args:
        stl_file_path (Path): The file path to the STL file.
        save_path (Path): The path where the generated preview image will be saved.
        image_size (tuple): The resolution of the output image as (width, height). Default is (200, 200).
        max_retries (int): The maximum number of retry attempts if an error occurs during rendering. 
                           Default is 10.
    
    Returns:
        bool: True if the preview was successfully created, False if all retries failed.
    
    Raises:
        ZeroDivisionError: If a division by zero occurs during rendering (handled internally).
        Exception: Any other unexpected errors during rendering are logged but do not raise exceptions.
    
    Notes:
        - The function applies color coding to the mesh based on the presence of '[a]' or '[c]' in the 
          file name: red for '[a]', white for '[c]', and black for all other files.
        - The camera parameters are set to provide an optimal view of the model, with auto-scaling 
          based on the model's dimensions.
        - If the image rendering fails due to a ZeroDivisionError (likely caused by a window size 
          calculation issue), the function retries up to 'max_retries' times.
    """
    retries = 0
    while retries < max_retries:
        try:
            mesh = trimesh.load_mesh(stl_file_path)
            scene = mesh.scene()

            # Apply color based on file name condition
            if '[a]' in stl_file_path.name.lower():
                mesh.visual.face_colors = [180, 0, 0, 255]  # Mild red
            elif '[c]' in stl_file_path.name.lower():
                mesh.visual.face_colors = [255, 255, 255, 255]  # White
            else:
                mesh.visual.face_colors = [50, 50, 50, 255]  # Mild black

            # Calculate optimal camera distance to fit the entire model
            bounding_box = mesh.bounding_box_oriented.bounds
            scale = max(bounding_box[1] - bounding_box[0])  # Size of the model's longest axis
            camera_distance = scale * 2.5  # Adjust zoom level based on model size

            # Ensure width and height are non-zero for aspect ratio calculation
            width, height = image_size
            if height == 0:
                height = 1

            # Set the camera parameters explicitly
            scene.set_camera(
                angles=[0.7, -0.3, 0.3],  # Front-top angle
                distance=camera_distance,
                center=mesh.centroid,  # Center the view on the model
            )

            # Set ambient lighting and background color for better contrast
            scene.ambient_light = [0.5, 0.5, 0.5, 1.0]  # Brighter ambient light
            scene.background = [240, 240, 240, 255]  # Light background

            # Save the image with suppressed preview display
            image = scene.save_image(resolution=image_size, visible=True)
            with open(save_path, 'wb') as f:
                f.write(image)
            
            return True
        except ZeroDivisionError as e:
            log_messages.append(f"ZeroDivisionError encountered when creating preview for {stl_file_path}: {e}")
            retries += 1
        except Exception as e:
            log_messages.append(f"Error creating preview for {stl_file_path}: {e}")
            return False
    # If all retries fail
    log_messages.append(f"Failed to create preview for {stl_file_path} after {max_retries} retries.")
    return False

# Initialize Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "STL Checklist"

# List to track missing previews
missing_previews = []

# Set headers
ws.append(["Filename", "Preview", "Checked and Available", "Not Needed"])

# Adjust column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15

# Make headers bold
for cell in ws[1]:
    cell.font = Font(bold=True)

# Track the current folder to add section headers
current_folder = None

# Process each STL file with a progress bar
for stl_path in tqdm(sorted(stl_files), desc="Processing STL files", unit="file"):
    # Determine the folder structure relative to 'Stls' directory
    folder = stl_path.parent.relative_to(stls_dir)
    
    # Add a new header for each folder if it changes
    if folder != current_folder:
        # Insert a blank row for spacing before new folder
        ws.append([""])
        
        # Add folder header row
        folder_row = ws.max_row + 1
        ws.append([f"Folder: {folder}"])
        ws[f"A{folder_row}"].font = Font(bold=True)
        
        current_folder = folder

    # Generate preview image path
    preview_path = temp_dir / f"{stl_path.name}.png"

    # Generate and save preview image
    if create_3d_preview(stl_path, preview_path):
        # Add filename and placeholders for checkboxes
        row = [
            stl_path.name,
            "",  # Placeholder for the image
            "",  # Placeholder for "Checked and Available"
            "",  # Placeholder for "Not Needed"
        ]
        ws.append(row)
        
        # Insert preview image into the Excel sheet
        img = Image(str(preview_path))
        img.width, img.height = 200, 200  # Resize image to 200x200 pixels
        img_cell = f"B{ws.max_row}"  # Column B, current row
        ws.add_image(img, img_cell)
        
        # Set the row height to 150 for rows with images
        ws.row_dimensions[ws.max_row].height = 150
    else:
        # Log missing preview and add entry to missing list with full path
        missing_info = f"{folder}/{stl_path.name}"
        missing_previews.append(missing_info)

        # Add a row in the main table even if preview is missing
        row = [
            stl_path.name,
            "Preview Missing",  # Indicate that the preview is missing
            "",  # Placeholder for "Checked and Available"
            "",  # Placeholder for "Not Needed"
        ]
        ws.append(row)

        # Set the row height to 20 for rows without images
        ws.row_dimensions[ws.max_row].height = 20

# Insert missing previews information at the top of the sheet if any previews are missing
if missing_previews:
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws["A1"] = "Warning: The following STL files failed to generate a preview:"
    ws["A1"].font = Font(bold=True, color="FF0000")
    for idx, missing_file in enumerate(missing_previews, start=2):
        ws[f"A{idx}"] = missing_file
    log_messages.append(f"{len(missing_previews)} STL previews failed to generate.")

# Save the Excel file
excel_output_path = Path.cwd() / "STL_Checklist_Structured.xlsx"
wb.save(excel_output_path)
log_messages.append(f"Checklist successfully saved at: {excel_output_path}")

# Clean up the temporary directory
shutil.rmtree(temp_dir)
log_messages.append(f"Temporary previews deleted from: {temp_dir}")

# Write the log messages to a log file with a timestamp
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_file_path = logs_dir / f"PRINTCHECK_log_{timestamp}.txt"
with open(log_file_path, 'w') as log_file:
    log_file.write("\n".join(log_messages))

# Display summary to the user
if missing_previews:
    print(f"{len(missing_previews)} STL previews could not be created. See {log_file_path} for details.")
else:
    print(f"All STL previews were created successfully. See {log_file_path} for details.")
